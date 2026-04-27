from __future__ import annotations

from io import BytesIO
from statistics import mean

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.core.logging import get_logger
from app.models.all_flight_result import AllFlightResult
from app.models.route_group import RouteGroup

log = get_logger(__name__)

_MAIN_HEADERS = [
    "Date",
    "Dep Airport",
    "Arrival Airport",
    "Nights",
    "Airline",
    "Flight Price",
]

_DEALS_HEADERS = [
    "Rank",
    "Origin",
    "Destination",
    "Date",
    "Airline",
    "Price",
    "Savings vs Avg",
]

_SUMMARY_HEADERS = [
    "Origin",
    "Records",
    "Lowest Price",
    "Average Price",
]

_WEEKEND_HEADERS = [
    "Origin",
    "Destination",
    "Date",
    "Airline",
    "Price",
]


def export_route_group(
    route_group: RouteGroup,
    all_results: list[AllFlightResult],
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    if not all_results:
        ws = wb.create_sheet("No Data")
        ws["A1"] = "No results available"
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    # --------------------------------------------------
    # LOOKUPS
    # --------------------------------------------------

    all_dates = sorted({r.depart_date for r in all_results})

    cheapest_by_origin_date: dict[tuple[str, object], AllFlightResult] = {}
    prices_by_route: dict[tuple[str, str], list[float]] = {}

    for r in all_results:
        key = (r.origin, r.depart_date)

        if key not in cheapest_by_origin_date:
            cheapest_by_origin_date[key] = r
        elif r.price < cheapest_by_origin_date[key].price:
            cheapest_by_origin_date[key] = r

        route_key = (r.origin, r.destination)
        prices_by_route.setdefault(route_key, []).append(float(r.price))

    # --------------------------------------------------
    # MAIN ORIGIN SHEETS
    # --------------------------------------------------

    sheet_name_map = route_group.sheet_name_map or {
        o: o for o in route_group.origins
    }

    for origin, sheet_name in sheet_name_map.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        _write_header_row(ws, _MAIN_HEADERS)

        for row_idx, d in enumerate(all_dates, start=2):
            result = cheapest_by_origin_date.get((origin, d))

            ws.cell(row=row_idx, column=1, value=d).number_format = "YYYY-MM-DD"
            ws.cell(row=row_idx, column=2, value=origin)
            ws.cell(
                row=row_idx,
                column=3,
                value=route_group.destination_label,
            )
            ws.cell(row=row_idx, column=4, value=route_group.nights)

            if result:
                ws.cell(row=row_idx, column=5, value=result.airline)
                ws.cell(
                    row=row_idx,
                    column=6,
                    value=int(round(float(result.price))),
                )

        _autosize_columns(ws)

    # --------------------------------------------------
    # SPECIAL JOURNEY SHEETS  (additional return / multi-city legs the
    # operator added in the "Advanced Routes" form)
    # --------------------------------------------------

    for sheet in route_group.special_sheets or []:
        sheet_name = (sheet.get("name") or "Journey")[:31]
        sheet_origin = (sheet.get("origin") or "").upper()
        sheet_dest_label = sheet.get("destination_label") or sheet_origin
        sheet_dests = [d.upper() for d in (sheet.get("destinations") or [])]
        columns = int(sheet.get("columns") or 4)

        ws = wb.create_sheet(title=sheet_name)

        if columns >= 6:
            _write_header_row(ws, _MAIN_HEADERS)
        else:
            _write_header_row(
                ws, ["Date", "Dep Airport", "Arrival Airport", "Flight Price"]
            )

        # cheapest result per date across this special sheet's destinations
        cheapest_per_date: dict[object, AllFlightResult] = {}
        for r in all_results:
            if r.origin != sheet_origin or r.destination not in sheet_dests:
                continue
            if (
                r.depart_date not in cheapest_per_date
                or r.price < cheapest_per_date[r.depart_date].price
            ):
                cheapest_per_date[r.depart_date] = r

        for row_idx, d in enumerate(all_dates, start=2):
            result = cheapest_per_date.get(d)

            ws.cell(row=row_idx, column=1, value=d).number_format = "YYYY-MM-DD"
            ws.cell(row=row_idx, column=2, value=sheet_origin)
            ws.cell(row=row_idx, column=3, value=sheet_dest_label)

            if columns >= 6:
                ws.cell(row=row_idx, column=4, value=route_group.nights)
                if result:
                    ws.cell(row=row_idx, column=5, value=result.airline)
                    ws.cell(
                        row=row_idx,
                        column=6,
                        value=int(round(float(result.price))),
                    )
            else:
                if result:
                    ws.cell(
                        row=row_idx,
                        column=4,
                        value=int(round(float(result.price))),
                    )

        _autosize_columns(ws)

    # --------------------------------------------------
    # BEST DEALS SHEET
    # --------------------------------------------------

    deals = []

    for r in all_results:
        route_key = (r.origin, r.destination)
        avg_price = mean(prices_by_route[route_key])

        savings = avg_price - float(r.price)

        deals.append(
            {
                "origin": r.origin,
                "destination": r.destination,
                "date": r.depart_date,
                "airline": r.airline,
                "price": float(r.price),
                "savings": savings,
            }
        )

    deals.sort(
        key=lambda x: (
            -x["savings"],
            x["price"],
        )
    )

    ws = wb.create_sheet("Best Deals")
    _write_header_row(ws, _DEALS_HEADERS)

    for i, d in enumerate(deals[:25], start=2):
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=d["origin"])
        ws.cell(row=i, column=3, value=d["destination"])
        ws.cell(row=i, column=4, value=d["date"]).number_format = "YYYY-MM-DD"
        ws.cell(row=i, column=5, value=d["airline"])
        ws.cell(row=i, column=6, value=int(round(d["price"])))
        ws.cell(row=i, column=7, value=int(round(d["savings"])))

    _autosize_columns(ws)

    # --------------------------------------------------
    # WEEKEND DEALS
    # --------------------------------------------------

    weekend = [
        r for r in all_results
        if r.depart_date.weekday() in (4, 5, 6)
    ]

    weekend.sort(key=lambda r: float(r.price))

    ws = wb.create_sheet("Weekend Deals")
    _write_header_row(ws, _WEEKEND_HEADERS)

    for i, r in enumerate(weekend[:25], start=2):
        ws.cell(row=i, column=1, value=r.origin)
        ws.cell(row=i, column=2, value=r.destination)
        ws.cell(row=i, column=3, value=r.depart_date).number_format = "YYYY-MM-DD"
        ws.cell(row=i, column=4, value=r.airline)
        ws.cell(row=i, column=5, value=int(round(float(r.price))))

    _autosize_columns(ws)

    # --------------------------------------------------
    # ORIGIN SUMMARY
    # --------------------------------------------------

    ws = wb.create_sheet("Summary")
    _write_header_row(ws, _SUMMARY_HEADERS)

    row = 2

    for origin in route_group.origins:
        rows = [r for r in all_results if r.origin == origin]

        if not rows:
            continue

        prices = [float(r.price) for r in rows]

        ws.cell(row=row, column=1, value=origin)
        ws.cell(row=row, column=2, value=len(rows))
        ws.cell(row=row, column=3, value=int(round(min(prices))))
        ws.cell(row=row, column=4, value=int(round(mean(prices))))

        row += 1

    _autosize_columns(ws)

    # --------------------------------------------------
    # FINISH
    # --------------------------------------------------

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.read()


def _write_header_row(ws, headers: list[str]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


def _autosize_columns(ws) -> None:
    for col_cells in ws.columns:
        max_length = max(
            (
                len(str(c.value))
                for c in col_cells
                if c.value is not None
            ),
            default=0,
        )

        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max_length + 3